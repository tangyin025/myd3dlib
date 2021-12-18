#!/usr/bin/python
# -*- coding: UTF-8 -*-
import os
import re
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer

dir="../tables_xlsx"
files=os.listdir(dir)
for file in files:
    path=dir+"/"+file
    if os.path.isfile(path):
        m=re.match("^([^~$].+)\.xlsx",file)
        if m:
            module=m.group(1)
            wb=load_workbook("../tables_xlsx/"+module+".xlsx",read_only=True)
            ws=wb.active
            o=open("../tables/"+module+".lua","w",encoding='utf-8')
            o.write("module(\""+module+"\", package.seeall)\n")
            o.write("data={\n")
            for row in ws.iter_rows(min_row=2,min_col=2,values_only=False):
                o.write("{")
                i=0
                for cell in row:
                    if i>0:
                        o.write(", ")
                    i=i+1
                    if cell.data_type=='n' and cell.value!=None:
                        o.write(str(cell.value))
                    elif cell.data_type=='s':
                        o.write('"'+cell.value.strip('"')+'"')
                    # elif cell.data_type=='f':
                    #     tok=Tokenizer(cell.value)
                    #     print("\n".join("%12s%11s%9s" % (t.value, t.type, t.subtype) for t in tok.items))
                    else:
                        o.write('nil')
                o.write("},\n")
            o.write("}\n")
            o.close()
            print(path)